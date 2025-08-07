from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
import sqlite3
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
from functools import wraps
import os
from pathlib import Path
import shutil
import io
import xlsxwriter
from openpyxl import Workbook, load_workbook
from contextlib import contextmanager
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from threading import Thread
from flask_mail import Mail, Message
from werkzeug.utils import secure_filename
import os

# Initialize Flask app
app = Flask(__name__)
app.secret_key = 'your_secure_secret_key_here'

# Email Configuration
app.config['MAIL_SERVER'] = 'smtp.example.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'your_email@example.com'
app.config['MAIL_PASSWORD'] = 'your_email_password'
app.config['MAIL_DEFAULT_SENDER'] = 'inventory@yourdomain.com'
app.config['ADMINS'] = ['admin@yourdomain.com']
app.config['LOW_STOCK_THRESHOLD'] = 10
app.config['EXPIRING_SOON_DAYS'] = 30

mail = Mail(app)

# Database Configuration
DATA_DIR = Path.home() / "frozen_management_data"
DB_PATH = DATA_DIR / "frozen.db"
EXCEL_LOG_PATH = DATA_DIR / "movements_log.xlsx"
BACKUP_DIR = DATA_DIR / "backups"
LOGO_PATH = Path(__file__).parent / "static" / "img" / "logo.png"

# Ensure directories exist
DATA_DIR.mkdir(exist_ok=True)
BACKUP_DIR.mkdir(exist_ok=True)

# Context processor to make current year available in templates
@app.context_processor
def inject_now():
    return {'now': datetime.now()}

# Database Context Manager
@contextmanager
def db_connection():
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    try:
        yield conn
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

# Initialize Database
def init_db():
    with db_connection() as conn:
        cursor = conn.cursor()
        
        # Create tables if they don't exist
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            role TEXT NOT NULL
        )''')
        
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS customers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE
        )''')
        
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            family TEXT NOT NULL,
            category TEXT NOT NULL
        )''')
        
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS movements (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER NOT NULL,
            quantity INTEGER NOT NULL,
            customer_id INTEGER,
            movement_type TEXT NOT NULL,
            date TEXT NOT NULL,
            best_before TEXT NOT NULL,
            batch TEXT NOT NULL,
            sub_batch TEXT NOT NULL,
            dpj TEXT NOT NULL,
            FOREIGN KEY (product_id) REFERENCES products (id),
            FOREIGN KEY (customer_id) REFERENCES customers (id)
        )''')
        
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS inventory (
            product_id INTEGER PRIMARY KEY,
            quantity INTEGER NOT NULL DEFAULT 0,
            FOREIGN KEY (product_id) REFERENCES products (id)
        )''')
        
        # Add this to your init_db() function
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS user_profiles (
            user_id INTEGER PRIMARY KEY,
            full_name TEXT,
            email TEXT UNIQUE,
            profile_picture TEXT,
            FOREIGN KEY (user_id) REFERENCES users (id)
        )''')
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS customers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            ville TEXT,
            pays TEXT,
            telephone TEXT,
            gsm TEXT,
            rc TEXT,
            cnss TEXT,
            patente TEXT,
            ice TEXT,
            observations TEXT
        )''')
        # Check if initial data needs to be added
        if conn.execute('SELECT COUNT(*) FROM users').fetchone()[0] == 0:
            # Initial admin user
            cursor.execute('''
            INSERT INTO users (username, password, role) 
            VALUES (?, ?, ?)''', ('admin', generate_password_hash('admin123'), 'admin'))
            
            # Sample products
            cursor.executemany('''
            INSERT INTO products (name, family, category) 
            VALUES (?, ?, ?)''', [
                ('Poulet', 'Poulet', 'Whole'),
                ('Dinde', 'Dinde', 'Whole')
            ])
            
            # Sample clients
            cursor.executemany('''
            INSERT INTO customers (name) 
            VALUES (?)''', [
                ('Restaurant ABC',),
                ('Hôtel XYZ',),
                ('Supermarket DEF',)
            ])
            
            # Initialize inventory
            cursor.execute('''
            INSERT INTO inventory (product_id, quantity)
            SELECT id, 0 FROM products
            ''')
        
        conn.commit()

# Excel Log Functions
def init_excel_log():
    if not EXCEL_LOG_PATH.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Mouvements"
        headers = [
            "ID", "Date", "Produit", "Famille", "Catégorie", "Type",
            "Quantité", "Client", "Lot", "Sous-lot", "DPJ", "DLC", "État"
        ]
        ws.append(headers)
        wb.save(str(EXCEL_LOG_PATH))

def update_excel_log(movement):
    try:
        # Create a temporary copy of the file
        temp_path = EXCEL_LOG_PATH.with_name(f"temp_{EXCEL_LOG_PATH.name}")
        
        # If Excel file is locked, create a new workbook
        if not EXCEL_LOG_PATH.exists() or not os.access(EXCEL_LOG_PATH, os.W_OK):
            wb = Workbook()
            ws = wb.active
            ws.title = "Mouvements"
            headers = [
                "ID", "Date", "Produit", "Famille", "Catégorie", "Type",
                "Quantité", "Client", "Lot", "Sous-lot", "DPJ", "DLC", "État"
            ]
            ws.append(headers)
        else:
            # Try to load the existing file
            try:
                wb = load_workbook(str(EXCEL_LOG_PATH))
                ws = wb.active
            except:
                # If loading fails, create new workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Mouvements"
                headers = [
                    "ID", "Date", "Produit", "Famille", "Catégorie", "Type",
                    "Quantité", "Client", "Lot", "Sous-lot", "DPJ", "DLC", "État"
                ]
                ws.append(headers)
        
        # Determine status
        status = "Bon"
        if get_alert_status(movement['best_before']) == 'danger':
            status = "Expiré"
        elif get_alert_status(movement['best_before']) == 'warning':
            status = "Bientôt"
        
        # Add the movement data
        ws.append([
        movement['id'],
        movement['date'][:16],
        movement.get('product_name', ''),
        movement.get('family', ''),
        movement.get('category', ''),
        movement['movement_type'],
        movement['quantity'],
        movement.get('customer_name', ''),
        movement['batch'],
        movement['sub_batch'],
        movement['dpj'],  # This will be the manually entered value
        movement['best_before'][:10],
        status
        ])

        # Save to temporary file first
        wb.save(str(temp_path))
        
        # Rename temporary file to original (atomic operation)
        if EXCEL_LOG_PATH.exists():
            os.remove(str(EXCEL_LOG_PATH))
        os.rename(str(temp_path), str(EXCEL_LOG_PATH))
        
    except Exception as e:
        print(f"Error updating Excel log: {str(e)}")

def calculate_dates(category, movement_type, product_name, dpj_date=None):
    # If dpj_date is provided, parse it, otherwise use current date
    if dpj_date:
        try:
            # Parse DD/MM/YYYY format
            day, month, year = map(int, dpj_date.split('/'))
            date = datetime(year, month, day)
        except ValueError:
            date = datetime.now()
    else:
        date = datetime.now()
    
    year_short = date.strftime('%y')
    
    # Calculate Julian day (1-365)
    julian_day = date.timetuple().tm_yday
    
    # Determine product code
    product_code = 'X'  # Default code
    if 'poulet' in product_name.lower():
        product_code = 'P'
    elif 'dinde' in product_name.lower():
        product_code = 'D'
    
    # Sub-batch format: AA-JJJ-C
    sub_batch = f"{year_short}-{julian_day:03d}-{product_code}"
    
    # Get current week number
    week_number = date.isocalendar()[1]
    
    # Batch format: AA+C+SS-SS2
    batch = f"{year_short}{product_code}{week_number}-{week_number+1}"
    
    # Calculate best before date based on category
    if category == 'MSM':
        bbd = date + timedelta(days=30*12)  # 12 months
    elif category == 'Offal':
        bbd = date + timedelta(days=30*9)  # 9 months
    else:  # Whole or Cut products (18 months)
        bbd = date + timedelta(days=30*18)  # 18 months
    
    return {
        'best_before': bbd.isoformat(),
        'batch': batch,
        'sub_batch': sub_batch,
    }

def get_alert_status(best_before_date):
    if isinstance(best_before_date, str):
        best_before_date = datetime.fromisoformat(best_before_date)
    
    delta = (best_before_date - datetime.now()).days
    
    if delta < 0:
        return 'danger'
    elif delta <= 30:
        return 'warning'
    else:
        return 'success'

def backup_db():
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = BACKUP_DIR / f"frozen_backup_{timestamp}.db"
    shutil.copy2(str(DB_PATH), str(backup_path))
    return str(backup_path)

# Email Functions
def send_async_email(app, msg):
    with app.app_context():
        mail.send(msg)

def send_email(subject, recipients, text_body, html_body=None):
    msg = Message(subject, recipients=recipients)
    msg.body = text_body
    if html_body:
        msg.html = html_body
    Thread(target=send_async_email, args=(app, msg)).start()

def check_inventory_alerts():
    with db_connection() as conn:
        # Low stock alert
        low_stock = conn.execute('''
        SELECT p.name, i.quantity 
        FROM inventory i
        JOIN products p ON i.product_id = p.id
        WHERE i.quantity < ?
        ''', (app.config['LOW_STOCK_THRESHOLD'],)).fetchall()
        
        if low_stock:
            subject = f"Low Stock Alert ({len(low_stock)} items)"
            text_body = "The following items are low on stock:\n\n" + \
                       "\n".join([f"{item['name']}: {item['quantity']} remaining" for item in low_stock])
            html_body = render_template('email/low_stock_alert.html', 
                                      items=low_stock,
                                      threshold=app.config['LOW_STOCK_THRESHOLD'])
            send_email(subject, app.config['ADMINS'], text_body, html_body)
        
        # Expiring soon alert
        expiring = conn.execute('''
        SELECT p.name, m.best_before, 
               julianday(m.best_before) - julianday('now') as days_left
        FROM movements m
        JOIN products p ON m.product_id = p.id
        WHERE m.best_before > datetime('now') 
          AND julianday(m.best_before) - julianday('now') <= ?
        GROUP BY m.product_id
        ''', (app.config['EXPIRING_SOON_DAYS'],)).fetchall()
        
        if expiring:
            subject = f"Expiring Products Alert ({len(expiring)} items)"
            text_body = "The following products will expire soon:\n\n" + \
                       "\n".join([f"{item['name']}: {item['best_before'][:10]} ({int(item['days_left'])} days)" for item in expiring])
            html_body = render_template('email/expiring_alert.html', 
                                      items=expiring,
                                      days=app.config['EXPIRING_SOON_DAYS'])
            send_email(subject, app.config['ADMINS'], text_body, html_body)

# Authentication Decorators
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in', 'danger')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session or session.get('role') != 'admin':
            flash('Admin access required', 'danger')
            return redirect(url_for('home'))
        return f(*args, **kwargs)
    return decorated_function

def manager_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session or session.get('role') not in ['admin', 'manager']:
            flash('Manager access required', 'danger')
            return redirect(url_for('home'))
        return f(*args, **kwargs)
    return decorated_function

# Initialize systems
with app.app_context():
    init_db()
    init_excel_log()

# Routes
@app.route('/')
@login_required
def home():
    return render_template('home.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        with db_connection() as conn:
            user = conn.execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()
        
        if user and check_password_hash(user['password'], password):
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user['role']
            flash('Login successful', 'success')
            return redirect(url_for('home'))
        else:
            flash('Invalid credentials', 'danger')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Logged out', 'info')
    return redirect(url_for('login'))

# Inventory Routes
@app.route('/inventory')
@login_required
def inventory_report():
    with db_connection() as conn:
        inventory = conn.execute('''
        SELECT p.id, p.name, p.family, p.category, 
               COALESCE(i.quantity, 0) as quantity
        FROM products p
        LEFT JOIN inventory i ON p.id = i.product_id
        ORDER BY p.name
        ''').fetchall()
    
    return render_template('inventory.html', inventory=inventory)

@app.route('/profile')
@login_required
def view_profile():
    with db_connection() as conn:
        profile = conn.execute('''
            SELECT u.id as user_id, u.username, u.role, p.full_name, p.email, p.profile_picture
            FROM users u
            LEFT JOIN user_profiles p ON u.id = p.user_id
            WHERE u.id = ?
        ''', (session['user_id'],)).fetchone()
    return render_template('profile.html', profile=profile, user_id=session['user_id'])

@app.route('/profile/edit', methods=['GET', 'POST'])
@login_required
def edit_profile():
    if request.method == 'POST':
        full_name = request.form.get('full_name')
        email = request.form.get('email')
        
        # Handle file upload
        profile_picture = None
        if 'profile_picture' in request.files:
            file = request.files['profile_picture']
            if file.filename != '':
                filename = secure_filename(file.filename)
                profile_pic_path = os.path.join(app.root_path, 'static', 'profile_pics', filename)
                file.save(profile_pic_path)
                profile_picture = f'profile_pics/{filename}'
        
        with db_connection() as conn:
            try:
                conn.execute('''
                    INSERT INTO user_profiles (user_id, full_name, email, profile_picture)
                    VALUES (?, ?, ?, ?)
                    ON CONFLICT(user_id) DO UPDATE SET
                    full_name = excluded.full_name,
                    email = excluded.email,
                    profile_picture = COALESCE(excluded.profile_picture, user_profiles.profile_picture)
                ''', (session['user_id'], full_name, email, profile_picture))
                conn.commit()
                flash('Profile updated successfully', 'success')
                return redirect(url_for('view_profile'))
            except sqlite3.IntegrityError:
                flash('Email already in use', 'danger')
    
    with db_connection() as conn:
        profile = conn.execute('''
            SELECT u.id as user_id, u.username, u.role, p.full_name, p.email, p.profile_picture
            FROM users u
            LEFT JOIN user_profiles p ON u.id = p.user_id
            WHERE u.id = ?
        ''', (session['user_id'],)).fetchone()
    return render_template('edit_profile.html', profile=profile)

@app.route('/profile/edit/<int:user_id>', methods=['GET', 'POST'])
@admin_required  # Only admins can access this route
def admin_edit_profile(user_id):
    if request.method == 'POST':
        full_name = request.form.get('full_name')
        email = request.form.get('email')
        role = request.form.get('role')
        
        # Handle file upload
        profile_picture = None
        if 'profile_picture' in request.files:
            file = request.files['profile_picture']
            if file.filename != '':
                filename = secure_filename(file.filename)
                profile_pic_path = os.path.join(app.root_path, 'static', 'profile_pics', filename)
                file.save(profile_pic_path)
                profile_picture = f'profile_pics/{filename}'
        
        with db_connection() as conn:
            try:
                # Update user profile
                conn.execute('''
                    INSERT INTO user_profiles (user_id, full_name, email, profile_picture)
                    VALUES (?, ?, ?, ?)
                    ON CONFLICT(user_id) DO UPDATE SET
                    full_name = excluded.full_name,
                    email = excluded.email,
                    profile_picture = COALESCE(excluded.profile_picture, user_profiles.profile_picture)
                ''', (user_id, full_name, email, profile_picture))
                
                # Update user role if changed (admin only)
                if role:
                    conn.execute('UPDATE users SET role = ? WHERE id = ?', (role, user_id))
                
                conn.commit()
                flash('Profile updated successfully', 'success')
                return redirect(url_for('admin_view_profile', user_id=user_id))
            except sqlite3.IntegrityError:
                flash('Email already in use', 'danger')
    
    with db_connection() as conn:
        profile = conn.execute('''
            SELECT u.id, u.username, u.role, p.full_name, p.email, p.profile_picture
            FROM users u
            LEFT JOIN user_profiles p ON u.id = p.user_id
            WHERE u.id = ?
        ''', (user_id,)).fetchone()
    return render_template('admin_edit_profile.html', profile=profile)

@app.route('/profile/view/<int:user_id>')
@manager_required
def admin_view_profile(user_id):
    with db_connection() as conn:
        profile = conn.execute('''
            SELECT u.id as user_id, u.username, u.role, p.full_name, p.email, p.profile_picture
            FROM users u
            LEFT JOIN user_profiles p ON u.id = p.user_id
            WHERE u.id = ?
        ''', (user_id,)).fetchone()
        
        if not profile:
            flash('User not found', 'danger')
            return redirect(url_for('manage_users'))
    
    return render_template('admin_view_profile.html', profile=profile, user_id=user_id)
    
    return render_template('admin_view_profile.html', profile=profile)
@app.route('/dashboard')
@manager_required
def dashboard():
    with db_connection() as conn:
        # Critical inventory
        critical = conn.execute('''
        SELECT p.name, p.family, p.category, i.quantity
        FROM inventory i
        JOIN products p ON i.product_id = p.id
        WHERE i.quantity < ?
        ORDER BY i.quantity ASC
        ''', (app.config['LOW_STOCK_THRESHOLD'],)).fetchall()
        
        # Expiring soon
        expiring = conn.execute('''
        SELECT p.name, p.family, p.category, 
               m.best_before, 
               julianday(m.best_before) - julianday('now') as days_left
        FROM movements m
        JOIN products p ON m.product_id = p.id
        WHERE m.best_before > datetime('now') 
          AND julianday(m.best_before) - julianday('now') <= ?
        GROUP BY m.product_id
        ORDER BY days_left ASC
        ''', (app.config['EXPIRING_SOON_DAYS'],)).fetchall()
        
        # Recent movements
        recent_movements = conn.execute('''
        SELECT m.date, p.name, m.quantity, m.movement_type, 
               c.name as customer_name
        FROM movements m
        JOIN products p ON m.product_id = p.id
        LEFT JOIN customers c ON m.customer_id = c.id
        ORDER BY m.date DESC
        LIMIT 10
        ''').fetchall()
        
        # Inventory summary
        inventory_summary = conn.execute('''
        SELECT 
            COUNT(CASE WHEN quantity = 0 THEN 1 END) as out_of_stock,
            COUNT(CASE WHEN quantity > 0 AND quantity < ? THEN 1 END) as low_stock,
            COUNT(CASE WHEN quantity >= ? THEN 1 END) as in_stock
        FROM inventory
        ''', (app.config['LOW_STOCK_THRESHOLD'], app.config['LOW_STOCK_THRESHOLD'])).fetchone()
    
    return render_template('dashboard.html',
                         critical=critical,
                         expiring=expiring,
                         recent_movements=recent_movements,
                         inventory_summary=inventory_summary,
                         threshold=app.config['LOW_STOCK_THRESHOLD'],
                         expiring_days=app.config['EXPIRING_SOON_DAYS'])

# Movement Routes
@app.route('/movements')
@login_required
def movements():
    product_filter = request.args.get('product', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    movement_type = request.args.get('movement_type', '')
    
    with db_connection() as conn:
        query = '''
        SELECT m.*, p.name as product_name, p.family, p.category, c.name as customer_name
        FROM movements m
        LEFT JOIN products p ON m.product_id = p.id
        LEFT JOIN customers c ON m.customer_id = c.id
        WHERE 1=1
        '''
        
        params = []
        
        if product_filter:
            query += ' AND p.name LIKE ?'
            params.append(f'%{product_filter}%')
            
        if date_from:
            query += ' AND date(m.date) >= ?'
            params.append(date_from)
            
        if date_to:
            query += ' AND date(m.date) <= ?'
            params.append(date_to)
            
        if movement_type:
            query += ' AND m.movement_type = ?'
            params.append(movement_type)
            
        query += ' ORDER BY m.date DESC'
        
        movements = conn.execute(query, params).fetchall()
        products = conn.execute('SELECT DISTINCT name FROM products ORDER BY name').fetchall()
    
    return render_template('movements.html', 
                         movements=movements, 
                         get_alert_status=get_alert_status,
                         products=products,
                         current_filters={
                             'product': product_filter,
                             'date_from': date_from,
                             'date_to': date_to,
                             'movement_type': movement_type
                         })

@app.route('/add_movement', methods=['GET', 'POST'])
@manager_required
def add_movement():
    if request.method == 'POST':
        try:
            with db_connection() as conn:
                product_id = request.form['product_id']
                quantity = int(request.form['quantity'])
                movement_type = request.form['movement_type']
                customer_id = request.form.get('customer_id') or None
                dpj = request.form['dpj']  # Get manually entered DPJ
                
                product = conn.execute('SELECT * FROM products WHERE id = ?', (product_id,)).fetchone()
                if not product:
                    flash('Product not found', 'danger')
                    return redirect(url_for('add_movement'))
                
                if movement_type == 'Exit':
                    current_stock = conn.execute('''
                    SELECT COALESCE(quantity, 0) as quantity 
                    FROM inventory WHERE product_id = ?
                    ''', (product_id,)).fetchone()
                    
                    if current_stock['quantity'] < quantity:
                        flash('Insufficient stock', 'danger')
                        return redirect(url_for('add_movement'))
                
                # Generate batch information using DPJ date
                dates = calculate_dates(product['category'], movement_type, product['name'], dpj)
                
                cursor = conn.execute('''
                INSERT INTO movements (
                    product_id, quantity, customer_id, movement_type, 
                    date, best_before, batch, sub_batch, dpj
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    product_id, quantity, customer_id, movement_type,
                    datetime.now().isoformat(), dates['best_before'], 
                    dates['batch'], dates['sub_batch'], dpj
                ))
                
                # Update inventory
                if movement_type == 'Entry':
                    conn.execute('''
                    INSERT INTO inventory (product_id, quantity)
                    VALUES (?, ?)
                    ON CONFLICT(product_id) DO UPDATE SET
                    quantity = quantity + excluded.quantity
                    ''', (product_id, quantity))
                else:
                    conn.execute('''
                    UPDATE inventory SET quantity = quantity - ?
                    WHERE product_id = ?
                    ''', (quantity, product_id))
                
                movement_id = cursor.lastrowid
                movement = conn.execute('''
                SELECT m.*, p.name as product_name, p.family, p.category, c.name as customer_name
                FROM movements m
                LEFT JOIN products p ON m.product_id = p.id
                LEFT JOIN customers c ON m.customer_id = c.id
                WHERE m.id = ?
                ''', (movement_id,)).fetchone()
                
                conn.commit()
                update_excel_log(dict(movement))
                check_inventory_alerts()
                flash('Movement recorded successfully', 'success')
                return redirect(url_for('view_receipt', movement_id=movement_id))
        
        except Exception as e:
            flash(f'Error: {str(e)}', 'danger')
            return redirect(url_for('add_movement'))
    
    with db_connection() as conn:
        products = conn.execute('SELECT * FROM products').fetchall()
        customers = conn.execute('SELECT * FROM customers ORDER BY name').fetchall()
    
    return render_template('add_movement.html', products=products, customers=customers)


@app.route('/receipt/<int:movement_id>')
@login_required
def view_receipt(movement_id):
    with db_connection() as conn:
        movement = conn.execute('''
        SELECT m.*, p.name as product_name, p.family, p.category, c.name as customer_name
        FROM movements m
        LEFT JOIN products p ON m.product_id = p.id
        LEFT JOIN customers c ON m.customer_id = c.id
        WHERE m.id = ?
        ''', (movement_id,)).fetchone()
    
    return render_template('movement_receipt.html', movement=movement)

@app.route('/print_receipt/<int:movement_id>')
@login_required
def print_receipt(movement_id):
    with db_connection() as conn:
        movement = conn.execute('''
        SELECT m.*, p.name as product_name, p.family, p.category, c.name as customer_name
        FROM movements m
        LEFT JOIN products p ON m.product_id = p.id
        LEFT JOIN customers c ON m.customer_id = c.id
        WHERE m.id = ?
        ''', (movement_id,)).fetchone()

    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, 
                          rightMargin=inch/2, leftMargin=inch/2,
                          topMargin=inch/2, bottomMargin=inch/2)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        alignment=1,
        spaceAfter=20,
        fontSize=16,
        textColor=colors.HexColor('#2c3e50')
    )
    header_style = ParagraphStyle(
        'Header',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.white,
        backColor=colors.HexColor('#4472C4'),
        spaceAfter=10
    )
    normal_style = styles['Normal']
    
    # Add logo if exists
    if LOGO_PATH.exists():
        logo = Image(str(LOGO_PATH), width=1.5*inch, height=0.75*inch)
        elements.append(logo)
    
    # Title and info
    elements.append(Paragraph(f"<b>CONDIFRI MAROC</b>", title_style))
    elements.append(Paragraph(f"Movement Receipt", title_style))
    elements.append(Spacer(1, 0.2*inch))
    elements.append(Paragraph(f"<b>Receipt #:</b> {movement['id']}", normal_style))
    elements.append(Paragraph(f"<b>Date:</b> {movement['date'][:16]}", normal_style))
    elements.append(Paragraph(f"<b>Type:</b> {movement['movement_type']}", normal_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Movement details
    elements.append(Paragraph("Movement Details", header_style))
    
    table_data = [
        ["Product", movement['product_name']],
        ["Family", movement['family']],
        ["Category", movement['category']],
        ["Quantity", movement['quantity']],
        ["Batch", movement['batch']],
        ["Sub-batch", movement['sub_batch']],
        ["DPJ", movement['dpj']],
        ["Best Before", movement['best_before'][:10]]
    ]
    
    details_table = Table(table_data, colWidths=[2*inch, 3*inch])
    details_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F5F5F5')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
    ]))
    elements.append(details_table)
    
    if movement['customer_name']:
        elements.append(Spacer(1, 0.2*inch))
        elements.append(Paragraph(f"<b>Customer:</b> {movement['customer_name']}", normal_style))
    
    # Footer
    elements.append(Spacer(1, 0.3*inch))
    elements.append(Paragraph("Condifri Maroc - Frozen Stock Management System", 
                            ParagraphStyle(name='Footer', alignment=1, fontSize=8)))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"movement_receipt_{movement['id']}.pdf",
        mimetype='application/pdf'
    )

# Client Routes
@app.route('/manage_clients')
@manager_required
def manage_clients():
    with db_connection() as conn:
        clients = conn.execute('SELECT * FROM customers ORDER BY name').fetchall()
        return render_template('clients.html', clients=clients)

@app.route('/add_client', methods=['GET', 'POST'])
@manager_required
def add_client():
    if request.method == 'POST':
        name = request.form['name']
        ville = request.form.get('ville', '')
        pays = request.form.get('pays', '')
        telephone = request.form.get('telephone', '')
        gsm = request.form.get('gsm', '')
        rc = request.form.get('rc', '')
        cnss = request.form.get('cnss', '')
        patente = request.form.get('patente', '')
        ice = request.form.get('ice', '')
        observations = request.form.get('observations', '')
        
        with db_connection() as conn:
            try:
                conn.execute('''
                INSERT INTO customers (
                    name, ville, pays, telephone, gsm, 
                    rc, cnss, patente, ice, observations
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    name, ville, pays, telephone, gsm, 
                    rc, cnss, patente, ice, observations
                ))
                conn.commit()
                flash('Client ajouté', 'success')
                return redirect(url_for('manage_clients'))
            except sqlite3.IntegrityError:
                flash('Client existe déjà', 'danger')
    
    return render_template('add_client.html')

@app.route('/edit_client/<int:client_id>', methods=['GET', 'POST'])
@manager_required
def edit_client(client_id):
    with db_connection() as conn:
        client = conn.execute('SELECT * FROM customers WHERE id = ?', (client_id,)).fetchone()
        
        if request.method == 'POST':
            name = request.form['name']
            ville = request.form.get('ville', '')
            pays = request.form.get('pays', '')
            telephone = request.form.get('telephone', '')
            gsm = request.form.get('gsm', '')
            rc = request.form.get('rc', '')
            cnss = request.form.get('cnss', '')
            patente = request.form.get('patente', '')
            ice = request.form.get('ice', '')
            observations = request.form.get('observations', '')
            
            try:
                conn.execute('''
                UPDATE customers SET 
                    name = ?, ville = ?, pays = ?, telephone = ?, gsm = ?,
                    rc = ?, cnss = ?, patente = ?, ice = ?, observations = ?
                WHERE id = ?
                ''', (
                    name, ville, pays, telephone, gsm,
                    rc, cnss, patente, ice, observations, client_id
                ))
                conn.commit()
                flash('Client modifié', 'success')
                return redirect(url_for('manage_clients'))
            except sqlite3.IntegrityError:
                flash('Nom déjà utilisé', 'danger')
        
        return render_template('edit_client.html', client=client)
    

@app.route('/delete_client/<int:client_id>', methods=['POST'])
@manager_required
def delete_client(client_id):
    with db_connection() as conn:
        try:
            conn.execute('DELETE FROM customers WHERE id = ?', (client_id,))
            conn.commit()
            flash('Client supprimé', 'success')
        except sqlite3.Error as e:
            flash(f'Erreur: {str(e)}', 'danger')
    return redirect(url_for('manage_clients'))

@app.route('/delete_movement/<int:movement_id>', methods=['POST'])
@manager_required
def delete_movement(movement_id):
    try:
        with db_connection() as conn:
            # Get movement details first
            movement = conn.execute('''
                SELECT * FROM movements WHERE id = ?
            ''', (movement_id,)).fetchone()
            
            if not movement:
                flash('Movement not found', 'danger')
                return redirect(url_for('movements'))
            
            # Update inventory (reverse the movement)
            if movement['movement_type'] == 'Entry':
                conn.execute('''
                    UPDATE inventory SET quantity = quantity - ?
                    WHERE product_id = ?
                ''', (movement['quantity'], movement['product_id']))
            else:  # Exit movement
                conn.execute('''
                    UPDATE inventory SET quantity = quantity + ?
                    WHERE product_id = ?
                ''', (movement['quantity'], movement['product_id']))
            
            # Delete the movement
            conn.execute('DELETE FROM movements WHERE id = ?', (movement_id,))
            conn.commit()
            
            flash('Movement deleted successfully', 'success')
    except Exception as e:
        flash(f'Error deleting movement: {str(e)}', 'danger')
    
    return redirect(url_for('movements'))

@app.route('/client/<int:client_id>')
@login_required
def client_details(client_id):
    with db_connection() as conn:
        client = conn.execute('SELECT * FROM customers WHERE id = ?', (client_id,)).fetchone()
        
        movements = conn.execute('''
        SELECT m.*, p.name as product_name, p.family, p.category
        FROM movements m
        JOIN products p ON m.product_id = p.id
        WHERE m.customer_id = ?
        ORDER BY m.date DESC
        ''', (client_id,)).fetchall()
        
        totals = conn.execute('''
        SELECT 
            SUM(CASE WHEN movement_type = 'Entry' THEN quantity ELSE 0 END) as total_entry,
            SUM(CASE WHEN movement_type = 'Exit' THEN quantity ELSE 0 END) as total_exit
        FROM movements WHERE customer_id = ?
        ''', (client_id,)).fetchone()
        
        return render_template('client_details.html',
                            client=client,
                            movements=movements,
                            total_entry=totals['total_entry'] or 0,
                            total_exit=totals['total_exit'] or 0,
                            get_alert_status=get_alert_status)

@app.route('/export_client_pdf/<int:client_id>')
@login_required
def export_client_pdf(client_id):
    with db_connection() as conn:
        client = conn.execute('SELECT * FROM customers WHERE id = ?', (client_id,)).fetchone()
        
        movements = conn.execute('''
        SELECT m.date, p.name as product_name, m.quantity, 
               m.movement_type, m.batch, m.sub_batch, m.dpj,
               m.best_before, p.family, p.category
        FROM movements m
        JOIN products p ON m.product_id = p.id
        WHERE m.customer_id = ?
        ORDER BY m.date DESC
        ''', (client_id,)).fetchall()
        
        totals = conn.execute('''
        SELECT 
            SUM(CASE WHEN movement_type = 'Entry' THEN quantity ELSE 0 END) as total_entry,
            SUM(CASE WHEN movement_type = 'Exit' THEN quantity ELSE 0 END) as total_exit
        FROM movements WHERE customer_id = ?
        ''', (client_id,)).fetchone()

    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, 
                          rightMargin=inch/2, leftMargin=inch/2,
                          topMargin=inch/2, bottomMargin=inch/2)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        alignment=1,
        spaceAfter=20,
        fontSize=16,
        textColor=colors.HexColor('#2c3e50')
    )
    header_style = ParagraphStyle(
        'Header',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.white,
        backColor=colors.HexColor('#4472C4'),
        spaceAfter=10
    )
    normal_style = styles['Normal']
    small_style = styles['Normal']
    small_style.fontSize = 8
    
    # Add logo if exists
    if LOGO_PATH.exists():
        logo = Image(str(LOGO_PATH), width=1.5*inch, height=0.75*inch)
        elements.append(logo)
    
    # Title and client info
    elements.append(Paragraph(f"<b>CONDIFRI MAROC</b>", title_style))
    elements.append(Paragraph(f"FICHE CLIENT", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Client Information Section
    elements.append(Paragraph("INFORMATIONS CLIENT", header_style))
    
    client_data = [
        ["Nom/Raison Sociale:", client['name']],
        ["Ville:", client['ville'] or '-'],
        ["Pays:", client['pays'] or '-'],
        ["Téléphone:", client['telephone'] or '-'],
        ["GSM:", client['gsm'] or '-'],
        ["RC:", client['rc'] or '-'],
        ["CNSS:", client['cnss'] or '-'],
        ["Patente:", client['patente'] or '-'],
        ["ICE:", client['ice'] or '-'],
        ["Observations:", client['observations'] or '-']
    ]
    
    client_table = Table(client_data, colWidths=[2*inch, 3*inch])
    client_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F5F5F5')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
    ]))
    elements.append(client_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Summary section
    elements.append(Paragraph("RÉSUMÉ DES MOUVEMENTS", header_style))
    
    summary_data = [
        ["Total Entrées:", f"{totals['total_entry'] or 0}"],
        ["Total Sorties:", f"{totals['total_exit'] or 0}"],
        ["Solde:", f"{(totals['total_entry'] or 0) - (totals['total_exit'] or 0)}"]
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 1*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F5F5F5')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Movements details
    if movements:
        elements.append(Paragraph("HISTORIQUE DES MOUVEMENTS", header_style))
        
        table_data = [
            ["Date", "Produit", "Type", "Qté", "Lot", "DLC", "État"]
        ]
        
        for movement in movements:
            status = get_alert_status(movement['best_before'])
            status_text = "Expiré" if status == 'danger' else "Bientôt" if status == 'warning' else "Bon"
            
            table_data.append([
                movement['date'][:16],
                movement['product_name'],
                movement['movement_type'],
                str(movement['quantity']),
                movement['batch'],
                movement['best_before'][:10],
                status_text
            ])
        
        movements_table = Table(table_data, colWidths=[1.2*inch, 1.5*inch, 0.6*inch, 0.5*inch, 0.8*inch, 0.8*inch, 0.8*inch])
        movements_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
        ]))
        elements.append(movements_table)
    else:
        elements.append(Paragraph("Aucun mouvement enregistré", normal_style))
    
    # Footer
    elements.append(Spacer(1, 0.3*inch))
    elements.append(Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}", 
                            ParagraphStyle(name='Footer', alignment=2, fontSize=8)))
    elements.append(Paragraph("Condifri Maroc - Système de Gestion de Stock Congelé", 
                            ParagraphStyle(name='Footer', alignment=1, fontSize=8)))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"fiche_client_{client['name']}_{datetime.now().strftime('%Y%m%d')}.pdf",
        mimetype='application/pdf'
    )

# Product Management Routes
@app.route('/manage_products')
@admin_required
def manage_products():
    with db_connection() as conn:
        products = conn.execute('SELECT * FROM products').fetchall()
        return render_template('products.html', products=products)

@app.route('/add_product', methods=['GET', 'POST'])
@admin_required
def add_product():
    if request.method == 'POST':
        name = request.form['name']
        family = request.form['family']
        category = request.form['category']
        
        with db_connection() as conn:
            try:
                conn.execute('''
                INSERT INTO products (name, family, category) 
                VALUES (?, ?, ?)
                ''', (name, family, category))
                conn.commit()
                flash('Produit ajouté', 'success')
            except sqlite3.Error as e:
                flash(f'Erreur: {str(e)}', 'danger')
            
            return redirect(url_for('manage_products'))
    
    return render_template('add_product.html')

@app.route('/edit_product/<int:product_id>', methods=['GET', 'POST'])
@admin_required
def edit_product(product_id):
    with db_connection() as conn:
        product = conn.execute('SELECT * FROM products WHERE id = ?', (product_id,)).fetchone()
        
        if request.method == 'POST':
            name = request.form['name']
            family = request.form['family']
            category = request.form['category']
            
            try:
                conn.execute('''
                UPDATE products SET name = ?, family = ?, category = ? WHERE id = ?
                ''', (name, family, category, product_id))
                conn.commit()
                flash('Produit modifié', 'success')
            except sqlite3.Error as e:
                flash(f'Erreur: {str(e)}', 'danger')
            
            return redirect(url_for('manage_products'))
        
        return render_template('edit_product.html', product=product)

@app.route('/delete_product/<int:product_id>', methods=['POST'])
@admin_required
def delete_product(product_id):
    with db_connection() as conn:
        try:
            conn.execute('DELETE FROM products WHERE id = ?', (product_id,))
            conn.commit()
            flash('Produit supprimé', 'success')
        except sqlite3.Error as e:
            flash(f'Erreur: {str(e)}', 'danger')
    return redirect(url_for('manage_products'))

# User Management Routes (Admin only)
@app.route('/manage_users')
@admin_required
def manage_users():
    with db_connection() as conn:
        users = conn.execute('SELECT * FROM users').fetchall()
        return render_template('users.html', users=users)

@app.route('/add_user', methods=['GET', 'POST'])
@admin_required
def add_user():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        role = request.form['role']
        
        with db_connection() as conn:
            try:
                conn.execute('''
                INSERT INTO users (username, password, role) 
                VALUES (?, ?, ?)
                ''', (username, generate_password_hash(password), role))
                conn.commit()
                flash('Utilisateur créé', 'success')
            except sqlite3.IntegrityError:
                flash('Nom d\'utilisateur existe déjà', 'danger')
            
            return redirect(url_for('manage_users'))
    
    return render_template('add_user.html')

@app.route('/edit_user/<int:user_id>', methods=['GET', 'POST'])
@admin_required
def edit_user(user_id):
    with db_connection() as conn:
        user = conn.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
        
        if request.method == 'POST':
            username = request.form['username']
            role = request.form['role']
            password = request.form.get('password')
            
            try:
                if password:
                    conn.execute('''
                    UPDATE users SET username = ?, password = ?, role = ? WHERE id = ?
                    ''', (username, generate_password_hash(password), role, user_id))
                else:
                    conn.execute('''
                    UPDATE users SET username = ?, role = ? WHERE id = ?
                    ''', (username, role, user_id))
                
                conn.commit()
                flash('Utilisateur modifié', 'success')
            except sqlite3.IntegrityError:
                flash('Nom d\'utilisateur existe déjà', 'danger')
            
            return redirect(url_for('manage_users'))
        
        return render_template('edit_user.html', user=user)

@app.route('/delete_user/<int:user_id>', methods=['POST'])
@admin_required
def delete_user(user_id):
    with db_connection() as conn:
        try:
            conn.execute('DELETE FROM users WHERE id = ?', (user_id,))
            conn.commit()
            flash('Utilisateur supprimé', 'success')
        except sqlite3.Error as e:
            flash(f'Erreur: {str(e)}', 'danger')
    return redirect(url_for('manage_users'))

# Database Maintenance
@app.route('/backup_db')
@admin_required
def create_backup():
    try:
        backup_path = backup_db()
        flash(f'Sauvegarde créée: {backup_path}', 'success')
    except Exception as e:
        flash(f'Erreur: {str(e)}', 'danger')
    return redirect(url_for('home'))

if __name__ == '__main__':
    app.run(debug=True)